import os
import pandas as pd
import pyodbc
import configparser
import openpyxl
from openpyxl.styles import PatternFill, Alignment

config = configparser.ConfigParser()
config.read('config.env')
db_UserName = config.get('DEFAULT', 'DB_USERNAME')
db_Password = config.get('DEFAULT', 'DB_PASSWORD')
db_Name = config.get('DEFAULT', 'DB_NAME')
db_Host = config.get('DEFAULT', 'DB_HOST')

cnxn_str = ("Driver={ODBC Driver 17 for SQL Server};"
            f"Server={db_Host};"
            f"Database={db_Name};"
            f"UID={db_UserName};"
            f"PWD={db_Password};")

cnxn = pyodbc.connect(cnxn_str)
# Create a cursor from the connection
cursor = cnxn.cursor()

# SQL查詢語句
query = ("select 　id,title ,context from ( "
         "select a.id,title,context from pttpost_referendum_4 a "
         "inner join pttpost b on a.source=b.source and a.id=b.Id "
         "where not exists (select * from keyword where source=99 and (b.title like '%'+keyname+'%' or b.context like '%'+keyname+'%')) "
         "union all "
         "select a.id,title,context from pttpost_referendum_4 a "
         "inner join pttpostgossing b on a.source=b.source and a.id=b.Id "
         "where not exists (select * from keyword where source=99 and (b.title like '%'+keyname+'%' or b.context like '%'+keyname+'%')) "
         "union all "
         "select convert(varchar,a.id),title,content from dcard.dbo.pttpost_referendum_4 a "
         "inner join dcard.dbo.post b on a.source=b.forum and a.id=b.Id "
         "where not exists (select * from keyword where source=99 and (b.title like '%'+keyname+'%' or b.content like '%'+keyname+'%')) "
         ") m "
         "where 1=1 "
         "and (m.title like N'%萊豬%' or m.context like N'%萊豬%') ")

cursor.execute(query)

# Export data to Excel
wb = openpyxl.Workbook()
ws = wb.active

# Write header row
header = ['標題', '內容']
ws.append(header)

# Write data rows
row_num = 2
datacnt = 0
for row in cursor.fetchall():
    title = row[1]
    context = row[2]
    combined = title + '\n' + context
    datacnt += 1
    ws.cell(row=row_num, column=1).value = str(datacnt)+"."
    for col in range(1, 30):
        ws.cell(row=row_num, column=col).fill = PatternFill(start_color='B0E0E6',
                                                            end_color='B0E0E6', fill_type='solid')

    # 合併儲存格
    ws.merge_cells(start_row=row_num+1, start_column=2,
                   end_row=row_num+1, end_column=12)
    # 設定儲存格格式及對齊方式
    context_cell = ws.cell(row=row_num+1, column=2)
    context_cell.value = combined
    context_cell.alignment = Alignment(wrap_text=True, vertical='top')
    context_cell.fill = PatternFill(start_color='FFD700',
                                    end_color='FFD700', fill_type='solid')

    ws.row_dimensions[row_num+1].height = 100  # 設定儲存格高度為100

    row_num += 2


# Save Excel file
wb.save('output4.xlsx')

# Close database connection
cursor.close()
cnxn.close()
